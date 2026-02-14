import os

from drf_excel.mixins import XLSXFileMixin
from drf_excel.renderers import XLSXRenderer
from rest_framework import serializers, status
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response


def header_map(values) -> dict[str, int]:
    return {
        str(value).strip(): idx
        for idx, value in enumerate(values)
        if value is not None
    }


class XlsxImportSerializer(serializers.Serializer):
    xlsx_file = serializers.FileField()


class XlsxExportImportMixin(XLSXFileMixin):
    export_serializer_class = None
    import_serializer_class = XlsxImportSerializer
    filename = "export.xlsx"

    def get_export_serializer_class(self):
        if self.export_serializer_class is None:
            raise AssertionError(
                "XlsxExportImportMixin requires `export_serializer_class` to be set."
            )
        return self.export_serializer_class

    def get_import_serializer_class(self):
        return self.import_serializer_class

    def get_serializer_class(self):
        if getattr(self, "action", None) == "import_xlsx":
            return self.get_import_serializer_class()
        return super().get_serializer_class()

    def get_import_required_headers(self):
        return []

    def get_import_expected_filename(self, request):
        return None

    def get_import_row_value(self, row, header_mapping, key):
        idx = header_mapping.get(key)
        return row[idx] if idx is not None else None

    def handle_import_row(self, row, header_mapping):
        raise NotImplementedError

    @action(
        detail=False,
        methods=["get"],
        permission_classes=[IsAdminUser],
        renderer_classes=[XLSXRenderer],
        url_path="export-xlsx",
    )
    def export_xlsx(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_export_serializer_class()(
            queryset,
            many=True,
            context={"request": request},
        )
        return Response(serializer.data)

    @action(
        detail=False,
        methods=["post"],
        permission_classes=[IsAdminUser],
        parser_classes=[MultiPartParser, FormParser],
        url_path="import-xlsx",
    )
    def import_xlsx(self, request):
        from openpyxl import load_workbook

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        upload = serializer.validated_data["xlsx_file"]

        try:
            expected_filename = self.get_import_expected_filename(request)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if expected_filename:
            actual_name = os.path.basename(upload.name or "")
            if actual_name != expected_filename:
                return Response(
                    {"detail": f"Filename must be {expected_filename}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        wb = load_workbook(upload)
        ws = wb.active
        header_mapping = header_map([cell.value for cell in ws[1]])
        required = self.get_import_required_headers()
        missing = [name for name in required if name not in header_mapping]
        if missing:
            return Response(
                {"detail": f"Missing columns: {', '.join(missing)}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            result = self.handle_import_row(row, header_mapping)
            if result == "created":
                created += 1
            elif result == "updated":
                updated += 1
            else:
                skipped += 1

        return Response({"created": created, "updated": updated, "skipped": skipped})
