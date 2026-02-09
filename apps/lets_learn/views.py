import io
from urllib.parse import urlparse
from urllib.request import urlopen

from django.conf import settings
from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.utils.text import slugify
from django.views.decorators.cache import cache_page
from openpyxl import Workbook, load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import BasePermission, SAFE_METHODS, IsAdminUser
from rest_framework.response import Response
from .models import CategoryConfig, LearnItem
from .serializers import CategorySerializer, LearnItemSerializer

class AdminWriteOrReadOnly(BasePermission):
    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:
            return True
        return bool(request.user and request.user.is_authenticated and request.user.is_staff)


def _download_to_content_file(url: str, fallback_name: str) -> tuple[ContentFile, str]:
    parsed = urlparse(url)
    filename = parsed.path.rsplit("/", 1)[-1] or fallback_name
    if "." not in filename:
        filename = fallback_name
    with urlopen(url) as response:
        content = response.read()
    return ContentFile(content), filename


def _header_map(values) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for idx, value in enumerate(values):
        if value is None:
            continue
        header_map[str(value).strip()] = idx
    return header_map


@method_decorator(cache_page(settings.CACHE_TTL), name="list")
@method_decorator(cache_page(settings.CACHE_TTL), name="retrieve")
class CategoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CategoryConfig.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [AdminWriteOrReadOnly]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        lang = (self.request.query_params.get('lang') or 'en').lower()
        if lang not in ('en', 'ne', 'hi'):
            lang = 'en'
        context['lang'] = lang
        return context

    @action(detail=False, methods=["get"], permission_classes=[IsAdminUser], url_path="export-xlsx")
    def export_xlsx(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Categories"
        ws.append(["category", "name", "slug", "image_url"])

        for category in self.get_queryset():
            image_url = (
                request.build_absolute_uri(category.image.url)
                if category.image
                else ""
            )
            ws.append([
                category.category,
                category.name,
                category.slug,
                image_url,
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = "attachment; filename=categories.xlsx"
        return response

    @action(detail=False, methods=["post"], permission_classes=[IsAdminUser], url_path="import-xlsx")
    def import_xlsx(self, request):
        upload = request.FILES.get("xlsx_file") or request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": "Missing XLSX file (field: xlsx_file or file)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        wb = load_workbook(upload)
        ws = wb.active
        header_map = _header_map([cell.value for cell in ws[1]])

        def get_value(row, key):
            idx = header_map.get(key)
            return row[idx] if idx is not None else None

        created = 0
        updated = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            category_value = get_value(row, "category")
            name = get_value(row, "name")
            image_url = get_value(row, "image_url")

            if not category_value:
                continue

            category_id = int(category_value)
            category, was_created = CategoryConfig.objects.get_or_create(
                category=category_id,
                defaults={"name": name or f"Category {category_id}"},
            )
            if was_created:
                created += 1
            else:
                updated += 1

            if name:
                category.name = name

            if image_url:
                fallback = f"category_{slugify(category.name) or category_id}.png"
                content, filename = _download_to_content_file(image_url, fallback)
                category.image.save(filename, content, save=False)

            category.save()

        return Response({"created": created, "updated": updated})

@method_decorator(cache_page(settings.CACHE_TTL), name="list")
@method_decorator(cache_page(settings.CACHE_TTL), name="retrieve")
class LearnItemViewSet(viewsets.ModelViewSet):
    queryset = LearnItem.objects.all()
    serializer_class = LearnItemSerializer
    filterset_fields = ['category']
    permission_classes = [AdminWriteOrReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        category_id = self.request.query_params.get('category')
        if category_id:
            queryset = queryset.filter(category=category_id)
        return queryset

    @action(detail=False, methods=["get"], permission_classes=[IsAdminUser], url_path="export-xlsx")
    def export_xlsx(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "LearnItems"
        ws.append(
            [
                "id",
                "category",
                "name",
                "content_name",
                "object_image_url",
                "object_color",
                "audio_url",
                "order",
            ]
        )

        for item in self.get_queryset():
            image_url = (
                request.build_absolute_uri(item.object_image.url)
                if item.object_image
                else ""
            )
            audio_url = (
                request.build_absolute_uri(item.audio.url)
                if item.audio
                else ""
            )
            ws.append(
                [
                    item.id,
                    item.category_id,
                    item.name,
                    item.content_name,
                    image_url,
                    item.object_color,
                    audio_url,
                    item.order,
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = "attachment; filename=learn_items.xlsx"
        return response

    @action(detail=False, methods=["post"], permission_classes=[IsAdminUser], url_path="import-xlsx")
    def import_xlsx(self, request):
        upload = request.FILES.get("xlsx_file") or request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": "Missing XLSX file (field: xlsx_file or file)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        wb = load_workbook(upload)
        ws = wb.active
        header_map = _header_map([cell.value for cell in ws[1]])

        def get_value(row, key):
            idx = header_map.get(key)
            return row[idx] if idx is not None else None

        created = 0
        updated = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_id = get_value(row, "id")
            category_value = get_value(row, "category")
            name = get_value(row, "name")
            content_name = get_value(row, "content_name")
            object_image_url = get_value(row, "object_image_url")
            object_color = get_value(row, "object_color")
            order = get_value(row, "order")

            if not category_value or not name:
                continue

            category = CategoryConfig.objects.filter(
                category=int(category_value)
            ).first()
            if not category:
                continue

            if item_id:
                item = LearnItem.objects.filter(id=int(item_id)).first()
            else:
                item = None

            if not item:
                item = LearnItem(category=category)
                created += 1
            else:
                updated += 1

            item.category = category
            item.name = name
            item.content_name = content_name
            if order is not None:
                item.order = int(order)

            if object_image_url:
                fallback = f"{slugify(name) or 'item'}.png"
                content, filename = _download_to_content_file(
                    object_image_url, fallback
                )
                item.object_image.save(filename, content, save=False)
                item.object_color = None
            elif object_color:
                item.object_color = object_color
                item.object_image = None

            item.save()

        return Response({"created": created, "updated": updated})
