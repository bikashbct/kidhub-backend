import io
from urllib.parse import urlparse
from urllib.request import urlopen

from django.contrib import admin
from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook

from .models import CategoryConfig, LearnItem


@admin.register(CategoryConfig)
class CategoryConfigAdmin(admin.ModelAdmin):
    list_display = ('category', 'name', 'slug')
    list_display_links = ('category', 'name')
    search_fields = ('name', 'slug')
    readonly_fields = ('slug',)
    change_list_template = "admin/lets_learn/categoryconfig/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-xlsx/",
                self.admin_site.admin_view(self.import_xlsx),
                name="lets_learn_categoryconfig_import_xlsx",
            ),
            path(
                "export-xlsx/",
                self.admin_site.admin_view(self.export_xlsx),
                name="lets_learn_categoryconfig_export_xlsx",
            ),
        ]
        return custom_urls + urls

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def export_xlsx(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Categories"
        ws.append(["category", "name", "slug", "image_url"])

        for category in self.get_queryset(request):
            image_url = (
                request.build_absolute_uri(category.image.url)
                if category.image
                else ""
            )
            ws.append([
                category.category,
                category.name,
                category.slug,
                image_url,
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = "attachment; filename=categories.xlsx"
        return response

    def import_xlsx(self, request):
        if request.method == "POST" and request.FILES.get("xlsx_file"):
            wb = load_workbook(request.FILES["xlsx_file"])
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}

            def get_value(row, key):
                idx = header_map.get(key)
                return row[idx] if idx is not None else None

            for row in ws.iter_rows(min_row=2, values_only=True):
                category_value = get_value(row, "category")
                name = get_value(row, "name")
                image_url = get_value(row, "image_url")

                if not category_value:
                    continue

                category_id = int(category_value)
                category, _ = CategoryConfig.objects.get_or_create(
                    category=category_id,
                    defaults={"name": name or f"Category {category_id}"},
                )

                if name:
                    category.name = name

                if image_url:
                    fallback = f"category_{slugify(category.name) or category_id}.png"
                    content, filename = _download_to_content_file(image_url, fallback)
                    category.image.save(filename, content, save=False)

                category.save()

            return redirect("..")

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Import categories from XLSX",
        }
        return render(request, "admin/lets_learn/categoryconfig/import_xlsx.html", context)

@admin.register(LearnItem)
class LearnItemAdmin(admin.ModelAdmin):
    list_display = ('name', 'slug', 'category_label', 'order')
    list_filter = ('category',)
    search_fields = ('name', 'slug')
    readonly_fields = ('slug',)
    change_list_template = "admin/lets_learn/learnitem/change_list.html"
    fields = (
        'name',
        'slug',
        'category',
        'content_name',
        'object_image',
        'object_color',
        'audio',
        'order',
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-xlsx/",
                self.admin_site.admin_view(self.import_xlsx),
                name="lets_learn_learnitem_import_xlsx",
            ),
            path(
                "export-xlsx/",
                self.admin_site.admin_view(self.export_xlsx),
                name="lets_learn_learnitem_export_xlsx",
            ),
        ]
        return custom_urls + urls

    def export_xlsx(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "LearnItems"
        ws.append(
            [
                "id",
                "category",
                "name",
                "content_name",
                "object_image_url",
                "object_color",
                "audio_url",
                "order",
            ]
        )

        for item in self.get_queryset(request):
            image_url = (
                request.build_absolute_uri(item.object_image.url)
                if item.object_image
                else ""
            )
            audio_url = (
                request.build_absolute_uri(item.audio.url)
                if item.audio
                else ""
            )
            ws.append(
                [
                    item.id,
                    item.category_id,
                    item.name,
                    item.content_name,
                    image_url,
                    item.object_color,
                    audio_url,
                    item.order,
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = "attachment; filename=learn_items.xlsx"
        return response

    def import_xlsx(self, request):
        if request.method == "POST" and request.FILES.get("xlsx_file"):
            wb = load_workbook(request.FILES["xlsx_file"])
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}

            def get_value(row, key):
                idx = header_map.get(key)
                return row[idx] if idx is not None else None

            for row in ws.iter_rows(min_row=2, values_only=True):
                item_id = get_value(row, "id")
                category_value = get_value(row, "category")
                name = get_value(row, "name")
                content_name = get_value(row, "content_name")
                object_image_url = get_value(row, "object_image_url")
                object_color = get_value(row, "object_color")
                order = get_value(row, "order")

                if not category_value or not name:
                    continue

                category = CategoryConfig.objects.filter(
                    category=int(category_value)
                ).first()
                if not category:
                    continue

                if item_id:
                    item = LearnItem.objects.filter(id=int(item_id)).first()
                else:
                    item = None

                if not item:
                    item = LearnItem(category=category)

                item.category = category
                item.name = name
                item.content_name = content_name
                if order is not None:
                    item.order = int(order)

                if object_image_url:
                    fallback = f"{slugify(name) or 'item'}.png"
                    content, filename = _download_to_content_file(
                        object_image_url, fallback
                    )
                    item.object_image.save(filename, content, save=False)
                    item.object_color = None
                elif object_color:
                    item.object_color = object_color

                item.save()

            return redirect("..")

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Import learn items from XLSX",
        }
        return render(request, "admin/lets_learn/learnitem/import_xlsx.html", context)

    @admin.display(description='Category')
    def category_label(self, obj):
        return obj.category.name if obj.category_id else None
